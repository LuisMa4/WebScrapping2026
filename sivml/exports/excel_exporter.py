from __future__ import annotations

import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from database import repository as repo
from database.models import Job, RawJob, ScrapingRun, Study


# ---------------------------------------------------------------------------
# Extraccion de requisitos desde texto de descripcion
# ---------------------------------------------------------------------------

_REQ_SECTION = re.compile(
    r"(?:\*{0,2}\s*)(?:requisitos?|requerimientos?|perfil\s+requerido?|"
    r"perfil\s+del\s+candidato?|what\s+we.re?\s+looking|qualifications?)"
    r"(?:\s*\*{0,2})?\s*:?\s*",
    re.IGNORECASE,
)
_SECTION_BREAK = re.compile(
    r"(?:funciones?|responsabilidades?|beneficios?|ofrecemos?|condiciones?"
    r"|que\s+ofrecemos?|acerca\s+de|sobre\s+(?:la|el)|salario|remuneraci"
    r"|condicion|beneficio|competencia|requerimiento\s+edu)",
    re.IGNORECASE,
)
_KW_REQUISITO = re.compile(
    r"(?:experiencia|estudios?|bachiller|titulado|egresado|licenciado"
    r"|conocimiento|manejo\s+de|dominio|habilidad|ingles|office|excel"
    r"|disponibilidad|residir|indispensable|deseable|certif)",
    re.IGNORECASE,
)


def _extract_requirements(description: str | None) -> str:
    """
    Extrae la seccion de requisitos del texto de la descripcion.
    Funciona con texto estructurado (con saltos de linea) y texto plano.
    """
    if not description:
        return ""

    # ── Estrategia 1: texto con saltos de linea (clean_description nuevo) ──
    if "\n" in description:
        lines = description.splitlines()
        in_req = False
        req_lines: list[str] = []

        for line in lines:
            stripped = line.strip()
            if not stripped:
                continue
            if _REQ_SECTION.search(stripped) and len(stripped) < 80:
                in_req = True
                continue
            if in_req:
                if _SECTION_BREAK.search(stripped) and len(stripped) < 80:
                    break
                req_lines.append(stripped.lstrip("-*•·✓►+ "))

        if req_lines:
            return "\n".join(l for l in req_lines if l)[:2000]

    # ── Estrategia 2: buscar la seccion en texto plano con regex ─────────────
    m = _REQ_SECTION.search(description)
    if m:
        after = description[m.end():]
        # Cortar en el proximo encabezado de seccion
        end = _SECTION_BREAK.search(after)
        req_text = after[: end.start()] if end else after[:2000]
        # Convertir los guiones inline en bullets de lista
        req_text = re.sub(r"\s+-\s*", "\n- ", req_text)
        req_text = re.sub(r"\s{2,}", " ", req_text).strip()
        if len(req_text) > 50:
            return req_text[:2000]

    # ── Estrategia 3: extraer lineas con palabras clave de requisito ─────────
    kw_lines = [
        line.strip().lstrip("-*•· ")
        for line in (description.split("\n") if "\n" in description
                     else re.split(r"(?<=[.!?])\s+", description))
        if _KW_REQUISITO.search(line) and len(line.strip()) > 15
    ]
    return "\n".join(kw_lines[:20])[:2000]

# ---------------------------------------------------------------------------
# Colores de cabecera
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_SUBHEADER_FILL = PatternFill("solid", fgColor="2E75B6")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_BOLD = Font(bold=True)


def _header_row(ws, values: list, fill=_HEADER_FILL):
    # ws.append() siempre añade en la siguiente fila disponible correctamente,
    # a diferencia de ws.max_row+1 que en hoja vacía devuelve fila 2 (bug openpyxl).
    ws.append(values)
    row = ws.max_row
    for col in range(1, len(values) + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _auto_width(ws, min_width: int = 10, max_width: int = 60):
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value or "")) for cell in col),
            default=min_width,
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            max(max_len + 2, min_width), max_width
        )


# ---------------------------------------------------------------------------
# Exportador principal
# ---------------------------------------------------------------------------

def export_study_to_excel(
    session: Session,
    study_id: str,
    output_dir: str | Path = "output",
) -> Path:
    study = repo.get_study(session, study_id)
    if study is None:
        raise ValueError(f"Estudio {study_id!r} no encontrado")

    jobs = repo.get_jobs_for_study(session, study_id)
    raw_jobs = repo.get_raw_jobs_for_study(session, study_id)
    runs = session.query(ScrapingRun).filter_by(study_id=study_id).all()

    wb = Workbook()
    wb.remove(wb.active)  # quitar hoja por defecto

    _sheet_resumen(wb, study, jobs, raw_jobs)
    _sheet_vacantes(wb, jobs)
    _sheet_vacantes_raw(wb, raw_jobs)
    _sheet_por_portal(wb, jobs, raw_jobs)
    _sheet_por_ciudad(wb, jobs)
    _sheet_por_empresa(wb, jobs)
    _sheet_tendencia(wb, jobs)
    _sheet_log_scraping(wb, runs)

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = output_dir / f"SIVML_{study_id[:8]}_{ts}.xlsx"
    wb.save(filename)
    return filename


# ---------------------------------------------------------------------------
# Hoja 1: Resumen
# ---------------------------------------------------------------------------

def _sheet_resumen(wb: Workbook, study: Study, jobs: Sequence[Job], raw_jobs: Sequence[RawJob]):
    ws = wb.create_sheet("Resumen")

    def kv(label, value):
        row = [label, value]
        ws.append(row)
        ws.cell(ws.max_row, 1).font = _BOLD

    ws.append(["SISTEMA INTELIGENTE DE VIGILANCIA DEL MERCADO LABORAL"])
    ws.cell(1, 1).font = Font(bold=True, size=14)
    ws.append([])
    kv("Estudio", study.name)
    kv("Programa académico", study.academic_program or "—")
    kv("ID del estudio", study.id)
    kv("Inicio", study.started_at.strftime("%Y-%m-%d %H:%M") if study.started_at else "—")
    kv("Fin", study.finished_at.strftime("%Y-%m-%d %H:%M") if study.finished_at else "—")
    kv("Estado", study.status)
    kv("Total scrapeadas", len(raw_jobs))
    kv("Únicas (post-dedup)", len(jobs))
    kv("Con salario", sum(1 for j in jobs if j.salary_min is not None))
    kv("Fecha generación", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC"))
    ws.append([])

    # Tabla por portal
    _header_row(ws, ["Portal", "Scrapeadas", "Nuevas", "Con salario"])
    portals: dict[str, dict] = defaultdict(lambda: {"scraped": 0, "new": 0, "salary": 0})
    for rj in raw_jobs:
        portals[rj.portal]["scraped"] += 1
    for j in jobs:
        if j.portal:
            portals[j.portal]["salary"] += 1 if j.salary_min else 0

    for portal, data in sorted(portals.items()):
        ws.append([portal, data["scraped"], data.get("new", 0), data["salary"]])

    _auto_width(ws)


# ---------------------------------------------------------------------------
# Hoja 2: Vacantes normalizadas
# ---------------------------------------------------------------------------

_VACANTES_COLS = [
    "ID", "Título Normalizado", "Empresa", "Ciudad", "País",
    "Portal", "URL", "Fecha Publicación", "Fecha Scraping",
    "Modalidad", "Tipo Contrato", "Exp. Min (años)", "Exp. Max (años)",
    "Nivel Educativo", "Salario Mín", "Salario Máx", "Moneda", "Período Salarial",
    "Requisitos", "Descripción Completa", "Keyword",
]


def _sheet_vacantes(wb: Workbook, jobs: Sequence[Job]):
    ws = wb.create_sheet("Vacantes")
    _header_row(ws, _VACANTES_COLS)
    ws.freeze_panes = "A2"

    for j in jobs:
        desc = j.description_clean or ""
        requisitos = _extract_requirements(desc)
        ws.append([
            j.id,
            j.title_normalized,
            j.company_normalized,
            j.city_normalized,
            j.country,
            j.portal,
            j.url,
            j.posted_date,
            j.created_at.date() if j.created_at else None,
            j.modality,
            j.contract_type,
            j.experience_years_min,
            j.experience_years_max,
            j.education_level,
            j.salary_min,
            j.salary_max,
            j.salary_currency,
            j.salary_period,
            requisitos,
            desc[:3000] if desc else "",
            "",  # keyword — disponible en raw_jobs, se omite aqui
        ])
        # Activar wrap_text en columnas de texto largo
        row_num = ws.max_row
        for col_name in ("Requisitos", "Descripción Completa"):
            col_idx = _VACANTES_COLS.index(col_name) + 1
            ws.cell(row_num, col_idx).alignment = Alignment(wrap_text=True, vertical="top")

    ws.auto_filter.ref = ws.dimensions
    # Ancho fijo para columnas de texto largo
    req_col = get_column_letter(_VACANTES_COLS.index("Requisitos") + 1)
    desc_col = get_column_letter(_VACANTES_COLS.index("Descripción Completa") + 1)
    ws.column_dimensions[req_col].width = 60
    ws.column_dimensions[desc_col].width = 80
    ws.row_dimensions[1].height = 30
    _auto_width(ws)


# ---------------------------------------------------------------------------
# Hoja 3: Vacantes crudas
# ---------------------------------------------------------------------------

_RAW_COLS = [
    "ID", "Portal", "Source ID", "Título Crudo", "Empresa Cruda", "Ciudad Cruda",
    "Fecha Publicación", "Fecha Scraping", "Salario Crudo", "Modalidad Cruda",
    "Contrato Crudo", "Experiencia Cruda", "Educación Cruda",
    "¿Duplicado?", "ID Canónico", "Keyword",
]


def _sheet_vacantes_raw(wb: Workbook, raw_jobs: Sequence[RawJob]):
    ws = wb.create_sheet("Vacantes_Raw")
    _header_row(ws, _RAW_COLS)
    ws.freeze_panes = "A2"

    for rj in raw_jobs:
        ws.append([
            rj.id,
            rj.portal,
            rj.source_id,
            rj.title,
            rj.company,
            rj.city,
            rj.posted_date,
            rj.scraped_at.date() if rj.scraped_at else None,
            rj.salary_raw,
            rj.modality_raw,
            rj.contract_raw,
            rj.experience_raw,
            rj.education_raw,
            "Sí" if rj.is_duplicate else "No",
            rj.canonical_id,
            rj.keyword_matched,
        ])

    ws.auto_filter.ref = ws.dimensions
    _auto_width(ws)


# ---------------------------------------------------------------------------
# Hoja 4: Por Portal (nueva)
# ---------------------------------------------------------------------------

def _sheet_por_portal(wb: Workbook, jobs: Sequence[Job], raw_jobs: Sequence[RawJob]):
    ws = wb.create_sheet("Por_Portal")
    _header_row(ws, [
        "Portal", "Vacantes Unicas", "% del Total",
        "Scrapeadas (raw)", "Salario Prom. Min", "Con Salario",
        "Con Empresa", "Con Ciudad",
    ])
    ws.freeze_panes = "A2"

    total_jobs = len(jobs) or 1
    total_raw = len(raw_jobs) or 1

    portal_jobs: dict[str, list[Job]] = defaultdict(list)
    for j in jobs:
        portal_jobs[j.portal or "desconocido"].append(j)

    portal_raw: dict[str, int] = defaultdict(int)
    for rj in raw_jobs:
        portal_raw[rj.portal] += 1

    for portal in sorted(portal_jobs, key=lambda p: -len(portal_jobs[p])):
        pjobs = portal_jobs[portal]
        salaries = [j.salary_min for j in pjobs if j.salary_min]
        avg_sal = round(sum(salaries) / len(salaries), 0) if salaries else None
        with_company = sum(1 for j in pjobs if j.company_normalized)
        with_city = sum(1 for j in pjobs if j.city_normalized)
        ws.append([
            portal,
            len(pjobs),
            round(len(pjobs) / total_jobs * 100, 1),
            portal_raw.get(portal, 0),
            avg_sal,
            len(salaries),
            with_company,
            with_city,
        ])

    _auto_width(ws)


# ---------------------------------------------------------------------------
# Hoja 5: Por ciudad
# ---------------------------------------------------------------------------

def _sheet_por_ciudad(wb: Workbook, jobs: Sequence[Job]):
    ws = wb.create_sheet("Por_Ciudad")
    _header_row(ws, ["Ciudad", "N Vacantes", "% del Total", "Salario Prom. Mín", "Salario Prom. Máx"])
    ws.freeze_panes = "A2"

    city_data: dict[str, list[Job]] = defaultdict(list)
    for j in jobs:
        city_data[j.city_normalized or "Desconocida"].append(j)

    total = len(jobs) or 1

    for city, city_jobs in sorted(city_data.items(), key=lambda x: -len(x[1])):
        salaries_min = [j.salary_min for j in city_jobs if j.salary_min]
        salaries_max = [j.salary_max for j in city_jobs if j.salary_max]
        avg_min = round(sum(salaries_min) / len(salaries_min), 0) if salaries_min else None
        avg_max = round(sum(salaries_max) / len(salaries_max), 0) if salaries_max else None
        ws.append([
            city,
            len(city_jobs),
            round(len(city_jobs) / total * 100, 1),
            avg_min,
            avg_max,
        ])

    _auto_width(ws)


# ---------------------------------------------------------------------------
# Hoja 5: Por empresa (top 50)
# ---------------------------------------------------------------------------

def _sheet_por_empresa(wb: Workbook, jobs: Sequence[Job]):
    ws = wb.create_sheet("Por_Empresa")
    _header_row(ws, ["Empresa", "N Vacantes", "Portales", "Salario Prom. Mín"])
    ws.freeze_panes = "A2"

    empresa_data: dict[str, list[Job]] = defaultdict(list)
    for j in jobs:
        empresa_data[j.company_normalized or "Desconocida"].append(j)

    top50 = sorted(empresa_data.items(), key=lambda x: -len(x[1]))[:50]

    for empresa, emp_jobs in top50:
        portales = ", ".join(sorted({j.portal or "" for j in emp_jobs if j.portal}))
        salaries = [j.salary_min for j in emp_jobs if j.salary_min]
        avg = round(sum(salaries) / len(salaries), 0) if salaries else None
        ws.append([empresa, len(emp_jobs), portales, avg])

    _auto_width(ws)


# ---------------------------------------------------------------------------
# Hoja 6: Tendencia temporal
# ---------------------------------------------------------------------------

def _sheet_tendencia(wb: Workbook, jobs: Sequence[Job]):
    ws = wb.create_sheet("Tendencia_Temporal")
    _header_row(ws, ["Año-Mes", "N Vacantes", "Acumulado"])
    ws.freeze_panes = "A2"

    monthly: dict[str, int] = defaultdict(int)
    for j in jobs:
        if j.posted_date:
            key = j.posted_date.strftime("%Y-%m")
            monthly[key] += 1

    acumulado = 0
    for period in sorted(monthly):
        acumulado += monthly[period]
        ws.append([period, monthly[period], acumulado])

    _auto_width(ws)


# ---------------------------------------------------------------------------
# Hoja 7: Log de scraping
# ---------------------------------------------------------------------------

def _sheet_log_scraping(wb: Workbook, runs: list[ScrapingRun]):
    ws = wb.create_sheet("Log_Scraping")
    _header_row(ws, [
        "ID", "Portal", "Keyword", "Ciudad",
        "Inicio", "Fin", "Encontradas", "Nuevas", "Estado", "Error",
    ])
    ws.freeze_panes = "A2"

    for run in runs:
        ws.append([
            run.id,
            run.portal,
            run.keyword,
            run.city,
            run.started_at,
            run.finished_at,
            run.records_found,
            run.records_new,
            run.status,
            run.error_message or "",
        ])

    _auto_width(ws)
