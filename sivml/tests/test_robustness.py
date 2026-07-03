"""
Tests de robustez: imports, portal_info, Excel, errores de portal, datos corruptos.
"""
import pytest
import sys
import os
from datetime import date
from pathlib import Path


# ---------------------------------------------------------------------------
# Import robustness
# ---------------------------------------------------------------------------

class TestImports:
    def test_portal_info_importable_without_scrapers(self):
        """portal_info.py no debe depender de ningun otro modulo del proyecto."""
        import importlib
        # Importar portal_info directamente (sin pasar por scrapers/__init__)
        spec = importlib.util.spec_from_file_location(
            "scrapers.portal_info",
            Path(__file__).parent.parent / "scrapers" / "portal_info.py",
        )
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        assert hasattr(mod, "PORTAL_STATUS")
        assert hasattr(mod, "PORTAL_CAPABILITIES")
        assert hasattr(mod, "RECOMMENDED_PORTALS")
        assert hasattr(mod, "SAFE_COMBINATIONS")

    def test_portal_status_always_available(self):
        from scrapers.portal_info import PORTAL_STATUS
        assert isinstance(PORTAL_STATUS, dict)
        assert len(PORTAL_STATUS) > 0

    def test_portal_status_via_scrapers_package(self):
        from scrapers import PORTAL_STATUS, RECOMMENDED_PORTALS, REGISTRY
        assert "computrabajo" in PORTAL_STATUS
        assert isinstance(RECOMMENDED_PORTALS, list)
        assert "computrabajo" in REGISTRY

    def test_get_scraper_unknown_portal(self):
        from scrapers import get_scraper
        with pytest.raises(ValueError, match="desconocido"):
            get_scraper("portal_que_no_existe")

    def test_all_portals_have_status(self):
        from scrapers import REGISTRY
        from scrapers.portal_info import PORTAL_STATUS
        for portal in REGISTRY:
            assert portal in PORTAL_STATUS, f"Portal '{portal}' sin entrada en PORTAL_STATUS"

    def test_all_portals_have_capabilities(self):
        from scrapers.portal_info import PORTAL_STATUS, PORTAL_CAPABILITIES
        for portal in PORTAL_STATUS:
            assert portal in PORTAL_CAPABILITIES, f"Portal '{portal}' sin entrada en PORTAL_CAPABILITIES"


# ---------------------------------------------------------------------------
# Portal capabilities integrity
# ---------------------------------------------------------------------------

class TestPortalCapabilities:
    REQUIRED_KEYS = [
        "uso_simultaneo", "nota_simultaneo", "max_keywords",
        "paginacion", "anti_bot", "delay_recomendado",
        "campos_disponibles", "requiere_login", "cobertura",
    ]

    def test_all_capability_keys_present(self):
        from scrapers.portal_info import PORTAL_CAPABILITIES
        for portal, cap in PORTAL_CAPABILITIES.items():
            for key in self.REQUIRED_KEYS:
                assert key in cap, f"Portal '{portal}' falta clave '{key}' en PORTAL_CAPABILITIES"

    def test_uso_simultaneo_is_bool(self):
        from scrapers.portal_info import PORTAL_CAPABILITIES
        for portal, cap in PORTAL_CAPABILITIES.items():
            assert isinstance(cap["uso_simultaneo"], bool), f"Portal '{portal}': uso_simultaneo debe ser bool"

    def test_campos_disponibles_is_list(self):
        from scrapers.portal_info import PORTAL_CAPABILITIES
        for portal, cap in PORTAL_CAPABILITIES.items():
            assert isinstance(cap["campos_disponibles"], list), f"Portal '{portal}': campos_disponibles debe ser list"

    def test_indeed_no_simultaneo(self):
        from scrapers.portal_info import PORTAL_CAPABILITIES
        assert PORTAL_CAPABILITIES["indeed"]["uso_simultaneo"] is False

    def test_computrabajo_simultaneo(self):
        from scrapers.portal_info import PORTAL_CAPABILITIES
        assert PORTAL_CAPABILITIES["computrabajo"]["uso_simultaneo"] is True

    def test_safe_combinations_only_include_known_portals(self):
        from scrapers.portal_info import PORTAL_STATUS, SAFE_COMBINATIONS
        known = set(PORTAL_STATUS.keys())
        for combo in SAFE_COMBINATIONS:
            for portal in combo:
                assert portal in known, f"Combo {combo} tiene portal desconocido '{portal}'"


# ---------------------------------------------------------------------------
# Excel export robustness
# ---------------------------------------------------------------------------

class TestExcelExport:
    @pytest.fixture
    def populated_session(self):
        """Session con un estudio y jobs para probar el exportador."""
        import tempfile
        from sqlalchemy import create_engine
        from sqlalchemy.orm import sessionmaker
        from database.session import Base
        import database.models  # registra modelos

        engine = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False})
        Base.metadata.create_all(engine)
        Session = sessionmaker(bind=engine)
        session = Session()

        from config.settings import StudyConfig, ScraperConfig
        from database import repository as repo
        from scrapers.base import ScrapedJob
        from datetime import datetime

        cfg = StudyConfig(
            study_id="excel-test-001",
            study_name="Excel Robustez Test",
            academic_program="Test",
            keywords=["sistemas"],
            cities=["Lima"],
            portals=["computrabajo", "bumeran"],
            date_from=date(2026, 1, 1),
            date_to=date(2026, 12, 31),
            scraper=ScraperConfig(),
        )
        study = repo.create_study(session, cfg)

        # Insertar raw_jobs simulados de dos portales
        for i in range(5):
            job = ScrapedJob(
                source_id=f"ct-{i}",
                portal="computrabajo",
                url=f"https://computrabajo.com/job/{i}",
                scraped_at=datetime.utcnow(),
                title=f"Ingeniero de Sistemas {i}",
                company="ACME SAC",
                city="Lima",
                country="Peru",
                salary_raw="S/ 3000",
                study_id="excel-test-001",
                keyword_matched="sistemas",
            )
            repo.upsert_raw_job(session, job)

        for i in range(3):
            job = ScrapedJob(
                source_id=f"bum-{i}",
                portal="bumeran",
                url=f"https://bumeran.com.pe/empleo/{i}",
                scraped_at=datetime.utcnow(),
                title=f"Analista de Sistemas {i}",
                company="Beta SA",
                city="Lima",
                country="Peru",
                study_id="excel-test-001",
                keyword_matched="sistemas",
            )
            repo.upsert_raw_job(session, job)

        from processing.deduplicator import run_exact_dedup
        run_exact_dedup(session, "excel-test-001")

        yield session, "excel-test-001"
        session.close()

    def test_excel_has_8_sheets(self, populated_session):
        import openpyxl, tempfile
        from exports.excel_exporter import export_study_to_excel
        session, study_id = populated_session
        output = Path(tempfile.mkdtemp())
        fp = export_study_to_excel(session, study_id, output_dir=output)
        wb = openpyxl.load_workbook(fp)
        expected = {"Resumen", "Vacantes", "Vacantes_Raw", "Por_Portal", "Por_Ciudad", "Por_Empresa", "Tendencia_Temporal", "Log_Scraping"}
        assert expected == set(wb.sheetnames), f"Hojas: {wb.sheetnames}"

    def test_por_portal_sheet_has_both_portals(self, populated_session):
        import openpyxl, tempfile
        from exports.excel_exporter import export_study_to_excel
        session, study_id = populated_session
        output = Path(tempfile.mkdtemp())
        fp = export_study_to_excel(session, study_id, output_dir=output)
        wb = openpyxl.load_workbook(fp)
        ws = wb["Por_Portal"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        pcol = headers.index("Portal") + 1
        portals = {ws.cell(r, pcol).value for r in range(2, ws.max_row + 1)}
        assert "computrabajo" in portals
        assert "bumeran" in portals

    def test_vacantes_sheet_headers_on_row_1(self, populated_session):
        import openpyxl, tempfile
        from exports.excel_exporter import export_study_to_excel
        session, study_id = populated_session
        output = Path(tempfile.mkdtemp())
        fp = export_study_to_excel(session, study_id, output_dir=output)
        wb = openpyxl.load_workbook(fp)
        ws = wb["Vacantes"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        assert "Portal" in headers, f"'Portal' no en fila 1. Headers: {headers}"
        assert "URL" in headers
        assert "Titulo Normalizado" in headers or "Título Normalizado" in headers

    def test_vacantes_portal_column_populated(self, populated_session):
        import openpyxl, tempfile
        from exports.excel_exporter import export_study_to_excel
        session, study_id = populated_session
        output = Path(tempfile.mkdtemp())
        fp = export_study_to_excel(session, study_id, output_dir=output)
        wb = openpyxl.load_workbook(fp)
        ws = wb["Vacantes"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        pcol = headers.index("Portal") + 1
        portals_in_data = {ws.cell(r, pcol).value for r in range(2, ws.max_row + 1)}
        assert "computrabajo" in portals_in_data
        assert "bumeran" in portals_in_data

    def test_excel_not_empty_when_no_jobs(self):
        """Exportar con 0 jobs no debe fallar — genera hojas vacias pero validas."""
        import openpyxl, tempfile
        from sqlalchemy import create_engine
        from sqlalchemy.orm import sessionmaker
        from database.session import Base
        import database.models
        from database import repository as repo
        from config.settings import StudyConfig, ScraperConfig
        from exports.excel_exporter import export_study_to_excel

        engine = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False})
        Base.metadata.create_all(engine)
        session = sessionmaker(bind=engine)()
        cfg = StudyConfig(
            study_id="empty-001",
            study_name="Empty Test",
            academic_program="Test",
            keywords=["x"],
            cities=["Lima"],
            portals=["computrabajo"],
            date_from=date(2026, 1, 1),
            date_to=date(2026, 12, 31),
            scraper=ScraperConfig(),
        )
        repo.create_study(session, cfg)
        output = Path(tempfile.mkdtemp())
        fp = export_study_to_excel(session, "empty-001", output_dir=output)
        assert fp.exists()
        assert fp.stat().st_size > 0
        wb = openpyxl.load_workbook(fp)
        assert "Vacantes" in wb.sheetnames
        session.close()


# ---------------------------------------------------------------------------
# Scraper robustness
# ---------------------------------------------------------------------------

class TestScraperRobustness:
    @pytest.fixture
    def cfg(self):
        from config.settings import StudyConfig, ScraperConfig
        return StudyConfig(
            study_id="rob-001",
            study_name="Robustez",
            academic_program="Test",
            keywords=["sistemas"],
            cities=["Lima"],
            portals=["computrabajo"],
            date_from=date(2026, 1, 1),
            date_to=date(2026, 12, 31),
            scraper=ScraperConfig(delay_range=(0.0, 0.0)),
        )

    def test_computrabajo_bad_html_returns_empty(self, cfg):
        """Scraper no debe lanzar excepcion con HTML malformado."""
        from scrapers.computrabajo import ComputrabajoScraper
        s = ComputrabajoScraper(cfg)
        result = s._parse_listing_page("<html><body>sin ofertas</body></html>")
        assert result == []

    def test_bumeran_no_cards_returns_empty(self, cfg):
        from scrapers.bumeran import BumeranScraper
        s = BumeranScraper(cfg)
        result = s._parse_listing_page("<html><body>vacio</body></html>")
        assert result == []

    def test_indeed_bad_html_returns_empty(self, cfg):
        from scrapers.indeed import IndeedScraper
        s = IndeedScraper(cfg)
        result = s._parse_listing_page("<html><body>nada</body></html>")
        assert result == []

    def test_normalizer_handles_none_inputs(self):
        from processing.normalizer import (
            normalize_city, normalize_company, normalize_salary,
            normalize_experience, normalize_education, normalize_modality, normalize_title,
        )
        assert normalize_city(None) is None
        assert normalize_company(None) is None
        assert normalize_salary(None) == (None, None, None, None)
        assert normalize_experience(None) == (None, None)
        assert normalize_education(None) is None
        assert normalize_modality(None) is None
        assert normalize_title(None) is None

    def test_normalizer_handles_empty_string(self):
        from processing.normalizer import normalize_city, normalize_salary
        assert normalize_city("") is None
        assert normalize_salary("") == (None, None, None, None)

    def test_normalizer_garbage_input(self):
        from processing.normalizer import normalize_salary, normalize_experience
        assert normalize_salary("abc xyz!!!") == (None, None, None, None)
        assert normalize_experience("no tiene experiencia") == (None, None)

    def test_deduplicator_empty_study(self):
        """Dedup en estudio vacio no debe fallar."""
        from sqlalchemy import create_engine
        from sqlalchemy.orm import sessionmaker
        from database.session import Base
        import database.models
        from database import repository as repo
        from config.settings import StudyConfig, ScraperConfig
        from processing.deduplicator import run_exact_dedup

        engine = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False})
        Base.metadata.create_all(engine)
        session = sessionmaker(bind=engine)()
        cfg = StudyConfig(
            study_id="empty-dedup-001",
            study_name="Empty Dedup",
            academic_program="Test",
            keywords=["x"],
            cities=["Lima"],
            portals=["computrabajo"],
            date_from=date(2026, 1, 1),
            date_to=date(2026, 12, 31),
            scraper=ScraperConfig(),
        )
        repo.create_study(session, cfg)
        stats = run_exact_dedup(session, "empty-dedup-001")
        assert stats["jobs_created"] == 0
        assert stats["duplicates_marked"] == 0
        session.close()
